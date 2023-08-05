# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author : Mike Zhou
# @Email : 公众号：测试开发技术
# @File : parse_excel.py

from openpyxl import load_workbook

class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        print(excelPath,sheetName)
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        dataList = []
        for line in self.sheet.rows:
            tmpList=[]
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            dataList.append(tmpList)
        return dataList[2:]
